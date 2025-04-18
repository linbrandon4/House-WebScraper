from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium import webdriver
from openpyxl import load_workbook
import numpy as np
import time
import re
import random

workBook = load_workbook("onMarketData.xlsx")
workSheet = workBook.active



options = Options()
service = Service("chromedriver.exe")
driver = webdriver.Chrome(service = service, options = options)

def increment_excel_column(column: str) -> str:
    def column_to_number(column):
        num = 0
        for char in column:
            num = num * 26 + (ord(char.upper()) - ord('A')) + 1
        return num
    
    def number_to_column(number):
        column_str = ""
        while number > 0:
            number, remainder = divmod(number - 1, 26)
            column_str = chr(remainder + ord('A')) + column_str
        return column_str
    
    column_num = column_to_number(column)
    incremented_column = number_to_column(column_num + 1)
    
    return incremented_column

def clickPopUp():
    try:
        driver.find_element(By.XPATH, '//*[@id="bp-dialog-container"]/div[1]/button/span').click()
    except:
        pass


baseURL = ["https://www.redfin.com/zipcode/30326/filter/property-type=house+condo+townhouse",
           "https://www.redfin.com/zipcode/30327/filter/property-type=house+condo+townhouse",
           "https://www.redfin.com/zipcode/30328/filter/property-type=house+condo+townhouse",
           "https://www.redfin.com/zipcode/30329/filter/property-type=house+condo+townhouse",
           "https://www.redfin.com/zipcode/30337/filter/property-type=house+condo+townhouse",
           "https://www.redfin.com/zipcode/30338/filter/property-type=house+condo+townhouse",
           "https://www.redfin.com/zipcode/30339/filter/property-type=house+condo+townhouse",
           "https://www.redfin.com/zipcode/30340/filter/property-type=house+condo+townhouse",
           "https://www.redfin.com/zipcode/30341/filter/property-type=house+condo+townhouse",
           "https://www.redfin.com/zipcode/30342/filter/property-type=house+condo+townhouse",
           "https://www.redfin.com/zipcode/30344/filter/property-type=house+condo+townhouse",
           "https://www.redfin.com/zipcode/30345/filter/property-type=house+condo+townhouse",
           "https://www.redfin.com/zipcode/30346/filter/property-type=house+condo+townhouse"]

totalPage = []
for urls in baseURL:
    allPages = []
    driver.get(urls)
    time.sleep(1)

    totalPages = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//span[@class='pageText']"))
    )
    total_pages_text = totalPages.text
    total_pages_number = int(total_pages_text.split()[-1])

    num = 2
    if total_pages_number == 1:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        pageLinks = WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.XPATH, "//a[@class='link-and-anchor visuallyHidden']"))
        )

        for page in pageLinks:
            href = page.get_attribute("href")
            if href:
                allPages.append(href)
    else:
        for x in range(total_pages_number - 1):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            pageLinks = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.XPATH, "//a[@class='link-and-anchor visuallyHidden']"))
            )
        
            for page in pageLinks:
                href = page.get_attribute("href")
                if href:
                    allPages.append(href)

            driver.get(urls + "/page-" + str(num))
            num += 1
            time.sleep(1)
    """
    zipCode = urls.split('/zipcode/')[1].split('/')[0]
    print("zip code: " + zipCode)

    empty_row = None
    for row_num in range(1, workS.max_row + 1):
        row_is_empty = all(cell.value is None or str(cell.value).strip() == "" for cell in workS[row_num])
        
        if row_is_empty:
            empty_row = row_num
            break

    if empty_row is None:
        empty_row = workS.max_row + 1

    zipCodeInSheet = [workS[f"A{row}"].value for row in range(2, workS.max_row + 1) if workS[f"A{row}"].value is not None]

    if zipCode not in zipCodeInSheet:
        empty_row = None
        for row_num in range(2, workS.max_row + 1):
            if workS[f'A{row_num}'].value is None or str(workS[f'A{row_num}'].value).strip() == "":
                empty_row = row_num
                break

        if empty_row is None:
            empty_row = workS.max_row + 1

        workS["A" + str(empty_row)] = zipCode
        workS["B" + str(empty_row)] = str(len(allPages))
        print("not in spreadsheet already " + zipCode)
    else:
        print("in spreadsheet already " + zipCode)
    """
    sample_size = int(0.2 * len(allPages))
    random_hrefs = random.sample(allPages, sample_size)

    totalPage += random_hrefs


urlnExcel = []
for row in workSheet.iter_rows(min_row = 2, max_col = 1, values_only = True):
    if row[0]:
        urlnExcel.append(row[0])

listOfURL = [url for url in totalPage if url not in urlnExcel]
print("total needed to scrape " + str(len(listOfURL)))

empty_row = None
for row_num in range(2, workSheet.max_row + 1):
    if workSheet[f'A{row_num}'].value is None:
        empty_row = row_num
        break

if empty_row is None:
    empty_row = workSheet.max_row + 1 

number = empty_row

for url in listOfURL:
    letter = "A"
    driver.get(url)
    time.sleep(1)

    workSheet[letter + str(number)].value = url
    workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        listedByRedfin = driver.find_element(By.XPATH, '//*[@data-rf-test-id="home-sash"]').text

        is_listed_by_redfin = "LISTED BY REDFIN" in listedByRedfin
        if is_listed_by_redfin:
            workSheet[letter + str(number)].value = "True"
            workBook.save("onMarketData.xlsx")
        else:
            workSheet[letter + str(number)].value = "False"
            workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "False"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        walkThrough = driver.find_element(By.XPATH, '//*[@aria-label="3D Walkthrough"]')
        workSheet[letter + str(number)].value = "True"
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "False"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        floorPlan = driver.find_element(By.XPATH, '//*[@aria-label="Floor Plans"]')
        workSheet[letter + str(number)].value = "True"
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "False"
        workBook.save("onMarketData.xlsx")
    clickPopUp()
    
    try:
        letter = increment_excel_column(letter)
        streetView = driver.find_element(By.XPATH, '//*[@aria-label="Street View"]')
        workSheet[letter + str(number)].value = "True"
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "False"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        redeisgn = driver.find_element(By.XPATH, '//*[@aria-label="Redesign"]')
        workSheet[letter + str(number)].value = "True"
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "False"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        photoCount = driver.find_element(By.XPATH, '//*[@id="photoPreviewButton"]/button/span[2]').text
        workSheet[letter + str(number)].value = photoCount
        workBook.save("onMarketData.xlsx")
        hasPhoto = True
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
        hasPhoto = False
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        address = driver.find_element(By.XPATH, '//*[@id="content"]/div[8]/div[2]/div[1]/div[1]/section/div/div/div/div[1]/div[2]/div/div/div/header/div/h1').text
        workSheet[letter + str(number)].value = address
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        price = driver.find_element(By.XPATH, '//*[@id="content"]/div[8]/div[2]/div[1]/div[1]/section/div/div/div/div[1]/div[2]/div/div/div/div/div[1]/div').text
        workSheet[letter + str(number)].value = price
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        beds = driver.find_element(By.XPATH, '//*[@id="content"]/div[8]/div[2]/div[1]/div[1]/section/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div').text
        workSheet[letter + str(number)].value = beds
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        baths = driver.find_element(By.XPATH, '//*[@id="content"]/div[8]/div[2]/div[1]/div[1]/section/div/div/div/div[1]/div[2]/div/div/div/div/div[3]/div').text
        workSheet[letter + str(number)].value = baths
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        sqft = driver.find_element(By.XPATH, '//*[@id="content"]/div[8]/div[2]/div[1]/div[1]/section/div/div/div/div[1]/div[2]/div/div/div/div/div[4]/span').text
        workSheet[letter + str(number)].value = sqft
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    time.sleep(1)

    try:
        letter = increment_excel_column(letter)
        estPayment = driver.find_element(By.XPATH, '//*[@id="content"]/div[8]/div[2]/div[1]/div[1]/section/div/div/div/div[1]/div[2]/div/div/div/div/div[1]/span/div/span[1]').text
        workSheet[letter + str(number)].value = estPayment
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        viewed = driver.find_element(By.XPATH, '//*[@id="content"]/div[8]/div[2]/div[1]/div[3]/section/div/div/section/div/div/div/div').text
        workSheet[letter + str(number)].value = viewed
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    time.sleep(1)
    try:
        letter = increment_excel_column(letter)
        button = driver.find_element(By.XPATH, '//*[@id="marketingRemarks-preview"]/div[2]/div/button').click()
        description = driver.find_element(By.XPATH, '//*[@id="marketingRemarks-preview"]/div[1] | //*[@id="marketing-remarks-scroll"] | //*[@id="marketing-remarks-scroll"]/p/span | //*[@id="marketing-remarks-scroll"]/p').text
        workSheet[letter + str(number)].value = description
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        ageOnRedfin = driver.find_element(By.XPATH, '//*[@id="house-info"]/div[2]/div/div/div/div/div[1]/div/div | //*[@id="house-info"]/div[3]/div/div/div/div/div[1]/div/div').text
        workSheet[letter + str(number)].value = ageOnRedfin
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        BuiltYear = driver.find_element(By.XPATH, '//*[@id="house-info"]/div[3]/div/div/div/div/div[3]/div | //*[@id="house-info"]/div[2]/div/div/div/div/div[3]/div').text
        workSheet[letter + str(number)].value = BuiltYear
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        style = driver.find_element(By.XPATH, '//*[@id="house-info"]/div[3]/div/div/div/div/div[2]/div | //*[@id="house-info"]/div[2]/div/div/div/div/div[2]/div').text
        workSheet[letter + str(number)].value = style
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        driver.find_element(By.XPATH, '//*[@id="house-info"]/div[3]/div/div/div[1]/div[2]/div/div[2]')
        listingAgent = driver.find_element(By.XPATH, '//*[@id="house-info"]/div[3]/div/div/div[1]/div[2]/div/div[2]/div/div/span[1]/a').text
        workSheet[letter + str(number)].value = listingAgent
        workBook.save("onMarketData.xlsx")

        letter = increment_excel_column(letter)
        listingAgency = driver.find_element(By.XPATH, '//*[@id="house-info"]/div[3]/div/div/div[1]/div[2]/div/div[2]/div/div/span[3]/span').text
        workSheet[letter + str(number)].value = listingAgency
        workBook.save("onMarketData.xlsx")
    except Exception as e:
        try:
            listing = driver.find_element(By.XPATH, '//*[@id="house-info"]/div[3]/div/div/div[1]/div[2]/div | //*[@id="house-info"]/div[4]/div/div/div[1]/div[2]/div').text
            parts = listing.splitlines()
        
            listingAgent = parts[0].strip() if len(parts) > 0 else None
            listingAgency = parts[1].strip() if len(parts) > 1 else None

            workSheet[letter + str(number)].value = listingAgent
            workBook.save("onMarketData.xlsx")

            letter = increment_excel_column(letter)
            workSheet[letter + str(number)].value = listingAgency
            workBook.save("onMarketData.xlsx")
        except:
            workSheet[letter + str(number)].value = "-"
            workBook.save("onMarketData.xlsx")
            
            letter = increment_excel_column(letter)
            workSheet[letter + str(number)].value = "-"
            workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        hasListing = False
        letter = increment_excel_column(letter)
        source = driver.find_element(By.XPATH, '//*[@id="house-info"]/div[3]/div/div/div[2]/div/div[2]/span[4] | //*[@id="house-info"]/div[4]/div/div/div[2]/div/div[2]/span[3] | //*[@id="house-info"]/div[3]/div/div/div[2]/div/div[2]/span[3]').text

        if source.strip().upper() == ("GAMLS"):
            workSheet[letter + str(number)].value = "GAMLS"
            workBook.save("onMarketData.xlsx")

            hasListing = True
            ssource = driver.find_element(By.XPATH, '//*[@id="house-info"]/div[4]/div/div/div[2]/div/div[2]/span[4] | //*[@id="house-info"]/div[3]/div/div/div[2]/div/div[2]/span[4]').text
            letter = increment_excel_column(letter)
            workSheet[letter + str(number)].value = ssource
            workBook.save("onMarketData.xlsx")
        
        else:
            workSheet[letter + str(number)].value = "FMLS"
            workBook.save("onMarketData.xlsx")
            letter = increment_excel_column(letter)

            workSheet[letter + str(number)].value = source
            workBook.save("onMarketData.xlsx")
    except:
        if hasListing:
            letter = increment_excel_column(letter)
            workSheet[letter + str(number)].value = "-"
            workBook.save("onMarketData.xlsx")
        else:
            workSheet[letter + str(number)].value = "-"
            workBook.save("onMarketData.xlsx")
            letter = increment_excel_column(letter)
            workSheet[letter + str(number)].value = "-"
            workBook.save("onMarketData.xlsx")
    
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        esName = driver.find_element(By.XPATH, '//*[@id="schools-scroll"]/div/div[2]/div[2]/div/div/div/div[1]/div/div/div/div | //*[@id="neighborhood-scroll"]/div/div[2]/div[2]/div/div/div/div[1]/div/div/div/div').text
        workSheet[letter + str(number)].value = esName
        workBook.save("onMarketData.xlsx")
    except Exception as e:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
        print(e)
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        esRating = driver.find_element(By.XPATH, '//*[@id="schools-scroll"]/div/div[2]/div[2]/div/div/div/div[1]/div/div/span[1]/div/span[1] | //*[@id="neighborhood-scroll"]/div/div[2]/div[2]/div/div/div/div[1]/div/div/span[1]/div/span[1]').text
        workSheet[letter + str(number)].value = esRating + "/10"
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        esDescription = driver.find_element(By.XPATH, '//*[@id="schools-scroll"]/div/div[2]/div[2]/div/div/div/div[1]/div/div/div/p | //*[@id="neighborhood-scroll"]/div/div[2]/div[2]/div/div/div/div[1]/div/div/div/p').text
        workSheet[letter + str(number)].value = esDescription
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        msName = driver.find_element(By.XPATH, '//*[@id="schools-scroll"]/div/div[2]/div[2]/div/div/div/div[2]/div/div/div/div | //*[@id="neighborhood-scroll"]/div/div[2]/div[2]/div/div/div/div[2]/div/div/div/div').text
        workSheet[letter + str(number)].value = msName
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        msRating = driver.find_element(By.XPATH, '//*[@id="schools-scroll"]/div/div[2]/div[2]/div/div/div/div[2]/div/div/span[1]/div/span[1] | //*[@id="neighborhood-scroll"]/div/div[2]/div[2]/div/div/div/div[2]/div/div/span[1]/div/span[1]').text
        workSheet[letter + str(number)].value = msRating + "/10"
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        msDescription = driver.find_element(By.XPATH, '//*[@id="schools-scroll"]/div/div[2]/div[2]/div/div/div/div[2]/div/div/div/p | //*[@id="neighborhood-scroll"]/div/div[2]/div[2]/div/div/div/div[2]/div/div/div/p').text
        workSheet[letter + str(number)].value = msDescription
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        hsName = driver.find_element(By.XPATH, '//*[@id="schools-scroll"]/div/div[2]/div[2]/div/div/div/div[3]/div/div/div/div | //*[@id="neighborhood-scroll"]/div/div[2]/div[2]/div/div/div/div[3]/div/div/div/div').text
        workSheet[letter + str(number)].value = hsName
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        hsRating = driver.find_element(By.XPATH, '//*[@id="schools-scroll"]/div/div[2]/div[2]/div/div/div/div[3]/div/div/span[1]/div/span[1] |//*[@id="neighborhood-scroll"]/div/div[2]/div[2]/div/div/div/div[3]/div/div/span[1]/div/span[1]').text
        workSheet[letter + str(number)].value = hsRating + "/10"
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()

    try:
        letter = increment_excel_column(letter)
        hsDescription = driver.find_element(By.XPATH, '//*[@id="schools-scroll"]/div/div[2]/div[2]/div/div/div/div[3]/div/div/div/p | //*[@id="neighborhood-scroll"]/div/div[2]/div[2]/div/div/div/div[3]/div/div/div/p').text
        workSheet[letter + str(number)].value  = hsDescription
        workBook.save("onMarketData.xlsx")
    except:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
    clickPopUp()
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

    time.sleep(1)

    try:
        letter = increment_excel_column(letter)
        allHouses = driver.find_elements(By.XPATH, '//a[contains(@href, "/home/")]')

        similarHrefs = [page.get_attribute('href') for page in allHouses]
        ssimilarHrefs = ', '.join(similarHrefs)

        workSheet[letter + str(number)].value  = ssimilarHrefs
        workBook.save("onMarketData.xlsx")
    except Exception as e:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")

    try:
        if hasPhoto:
            letter = increment_excel_column(letter)
            imgButton = driver.find_element(By.XPATH, '//*[@id="MBImage0"]/img')
            driver.execute_script("arguments[0].click();", imgButton)

            time.sleep(2)
            images = driver.find_elements(By.CLASS_NAME, 'img-card')

            srcList = [img.get_attribute('src') for img in images]
            srcString = ' '.join(srcList)

            workSheet[letter + str(number)].value = srcString
            workBook.save("onMarketData.xlsx")
        else:
            letter = increment_excel_column(letter)
            workSheet[letter + str(number)].value = "-"
            workBook.save("onMarketData.xlsx")
    except Exception as e:
        workSheet[letter + str(number)].value = "-"
        workBook.save("onMarketData.xlsx")
        print(e)

    number += 1
