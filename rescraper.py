from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium import webdriver
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime

# Replace with list of cities you are rescraping, make sure that the name matches your files
names = ["Nashville", "Los Angeles", "New York City", 
         "Oklahoma City", "San Jose", "Seattle"]

options = Options()
service = Service("chromedriver.exe")
driver = webdriver.Chrome(service=service, options=options)

def calculate_days(date1, date2):
    if date1 == "-" or date2 == "-":
        return "-"
    try:
        date_format = "%b %d, %Y"
        d1 = datetime.strptime(date1, date_format)
        d2 = datetime.strptime(date2, date_format)
        return (d2 - d1).days
    except Exception as e:
        print(f"Error calculating days: {e}")
        return "-"

def clickPopUp():
    try:
        driver.find_element(By.XPATH, '//*[@id="bp-dialog-container"]/div[1]/button/span').click()
    except:
        pass

for name in names:
    workBook = load_workbook(name + ".xlsx")
    workSheet = workBook.active

    # Ensure headers are set up (if needed):
    workSheet['AC1'].value = "Listing Price"
    workSheet['AD1'].value = "Sell Price"
    workSheet['AE1'].value = "Price Changed Count"
    workSheet['AF1'].value = "Listing Date"
    workSheet['AG1'].value = "Selling Date"
    workSheet['AH1'].value = "Pending Date"
    workSheet['AI1'].value = "Source"
    workSheet['AJ1'].value = "Listed to Pending Day"
    workSheet['AK1'].value = "Listed to Sold Days"
    workSheet['AL1'].value = "Listing Removed"
    workSheet['AM1'].value = "Listing Removal Date"
    workSheet['AN1'].value = "Rent"
    workSheet['AO1'].value = "All Sales History"
    workBook.save(name + ".xlsx")

    # Keywords
    list_keywords = ["Listed (Active)", "Listed"]
    sell_keywords = ["Sold"]
    pending_keywords = ["Pending"]
    price_changed_keyword = "Price Changed"
    removed_keywords = ["Listing Removed"]
    rent_keyword = "Listed for Rent"  # The exact phrase to detect rentals

    # Gather all URLs from column A (rows 2+)
    urls = []
    for row in workSheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            urls.append(row[0])

    # We will process only 10 URLs for each city
    num = 1  # We'll increment at the start of each loop

    for url in urls:
        # Increment num for this row index
        num += 1

        # If num > 11, we've processed rows 2..11 (10 total)


        # If AC already has a value, skip this row
        if workSheet['AC' + str(num)].value not in ("", None):
            continue

        driver.get(url)
        clickPopUp()

        # Try to see if it's "SOLD", "LAST SOLD", or "OFF MARKET"
        try:
            sold_text = driver.find_element(
                By.XPATH, 
                '//*[@id="content"]/div[8]/div[2]/div[1]/div[1]/section/div/div[1]/div/div[1]/div[1]/span'
            ).text
        except:
            sold_text = "not"

        # If not sold/off market, skip (do not fill anything in).
        if sold_text not in ["SOLD", "LAST SOLD", "OFF MARKET"]:
            continue

        # NEW OR MODIFIED: Attempt to find property history; if not found, skip.
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "PropertyHistoryEventRow"))
            )
        except:
            # If we cannot find the sales history at all, just skip this listing
            continue

        # Now parse the property history
        try:
            data = {
                "list_price": "-",
                "sell_price": "-",
                "price_changed_count": 0,
                "listing_date": "-",
                "selling_date": "-",
                "pending_date": "-",
                "listing_removed_date": "-",
                "most_recent_source": "-",
                "listed_to_pending_days": "-",
                "listed_to_sold_days": "-",
                "rent": False,
                "all_sales_history": ""
            }

            listing_removed_flag = False
            all_history_entries = []

            event_rows = driver.find_elements(By.CLASS_NAME, "PropertyHistoryEventRow")
            if not event_rows:
                # No rows present, skip
                continue

            for row_elem in event_rows:
                try:
                    description_element = row_elem.find_element(By.CSS_SELECTOR, ".description-col div")
                    description_text = description_element.text.strip()

                    date_element = row_elem.find_element(By.CSS_SELECTOR, ".col-4 p")
                    event_date = date_element.text.strip()

                    price_element = row_elem.find_element(By.CSS_SELECTOR, ".price-col.number")
                    price_text = price_element.text.strip()

                    source_element = row_elem.find_element(By.CSS_SELECTOR, ".description-col p.subtext")
                    source_text = source_element.text.strip()

                    # Set most recent source if empty
                    if data["most_recent_source"] == "-":
                        data["most_recent_source"] = source_text

                    # Collect the event info for "All Sales History"
                    all_history_entries.append(f"{event_date} - {description_text} - {price_text}")

                    # Check for rent keyword
                    if rent_keyword.lower() in description_text.lower():
                        data["rent"] = True

                    # Set listing date
                    if any(keyword in description_text for keyword in list_keywords) and data["listing_date"] == "-":
                        data["listing_date"] = event_date
                        data["list_price"] = price_text

                    # Set selling date
                    if any(keyword in description_text for keyword in sell_keywords) and data["selling_date"] == "-":
                        data["selling_date"] = event_date
                        data["sell_price"] = price_text

                    # Set pending date
                    if any(keyword in description_text for keyword in pending_keywords) and data["pending_date"] == "-":
                        data["pending_date"] = event_date

                    # Count price changes
                    if price_changed_keyword in description_text:
                        data["price_changed_count"] += 1

                    # Check if listing removed
                    if any(keyword in description_text for keyword in removed_keywords):
                        listing_removed_flag = True
                        if data["listing_removed_date"] == "-":
                            data["listing_removed_date"] = event_date
                        else:
                            # Keep the latest removal date
                            try:
                                current_removed_date = datetime.strptime(event_date, "%b %d, %Y")
                                existing_removed_date = datetime.strptime(data["listing_removed_date"], "%b %d, %Y")
                                if current_removed_date > existing_removed_date:
                                    data["listing_removed_date"] = event_date
                            except:
                                pass

                except Exception as e:
                    print(f"Error processing row event for {url}: {e}")

            # Join all history entries
            data["all_sales_history"] = "/".join(all_history_entries)

            # Calculate days from listing to pending/sold
            data["listed_to_pending_days"] = calculate_days(data["listing_date"], data["pending_date"])
            data["listed_to_sold_days"] = calculate_days(data["listing_date"], data["selling_date"])

            print(f"Property: {url}")
            print("List Price:", data["list_price"])
            print("Sell Price:", data["sell_price"])
            print("Price Changed Count:", data["price_changed_count"])
            print("Listing Date:", data["listing_date"])
            print("Selling Date:", data["selling_date"])
            print("Pending Date:", data["pending_date"])
            print("Most Recent Source:", data["most_recent_source"])
            print("Listed to Pending Days:", data["listed_to_pending_days"])
            print("Listed to Sold Days:", data["listed_to_sold_days"])
            if listing_removed_flag:
                print("Listing removed on this date:", data["listing_removed_date"])
            print("Was it ever listed for rent?:", data["rent"])
            print("All Sales History:", data["all_sales_history"])
            print("-" * 40)

            # Write data back to Excel
            workSheet['AC' + str(num)].value = data["list_price"]
            workSheet['AD' + str(num)].value = data["sell_price"]
            workSheet['AE' + str(num)].value = str(data["price_changed_count"])
            workSheet['AF' + str(num)].value = data["listing_date"]
            workSheet['AG' + str(num)].value = data["selling_date"]
            workSheet['AH' + str(num)].value = data["pending_date"]
            workSheet['AI' + str(num)].value = data["most_recent_source"]
            workSheet['AJ' + str(num)].value = str(data["listed_to_pending_days"])
            workSheet['AK' + str(num)].value = str(data["listed_to_sold_days"])
            workSheet['AL' + str(num)].value = listing_removed_flag
            workSheet['AM' + str(num)].value = data["listing_removed_date"]
            workSheet['AN' + str(num)].value = data["rent"]
            workSheet['AO' + str(num)].value = data["all_sales_history"]

            workBook.save(name + ".xlsx")

        except Exception as e:
            print(f"Error loading property history for {url}: {e}")

# Done with this city, move to the next city
driver.quit()
